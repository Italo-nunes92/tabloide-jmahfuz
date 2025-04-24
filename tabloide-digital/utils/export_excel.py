import openpyxl
from io import BytesIO
from django.http import HttpResponse

def export_as_excel(stats, headers):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chamadas no WhatsApp"
    
    # Preparar dados
    data = []
    for i, v in stats.items():
        for f in v:
            data.append([i, *f.values()])
    
    # Adicionar cabeçalhos
    ws.append(headers)
    
    # Adicionar dados
    for row in data:
        ws.append(row)
    
    # Definir estilos baseados no CSS
    federal_blue_fill = PatternFill(start_color="1A1A66", end_color="1A1A66", fill_type="solid")
    steel_blue_fill = PatternFill(start_color="3A85BD", end_color="3A85BD", fill_type="solid")
    
    # Estilo para cabeçalhos
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Bordas
    thin_border = Border(
        left=Side(style='thin', color="DDDDDD"),
        right=Side(style='thin', color="DDDDDD"),
        top=Side(style='thin', color="DDDDDD"),
        bottom=Side(style='thin', color="DDDDDD")
    )
    
    # Aplicar estilos aos cabeçalhos
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = steel_blue_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Aplicar estilos às células de dados
    for row_idx, row in enumerate(data, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Aplicar estilo zebrado
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Ajustar largura das colunas com base no conteúdo
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        # Encontrar o comprimento máximo em cada coluna
        max_length = 0
        for row_idx in range(1, len(data) + 2):  # +2 para incluir o cabeçalho e começar do 1
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                # Adicionar um pequeno buffer para espaçamento
                cell_length = len(str(cell_value)) + 2
                max_length = max(max_length, cell_length)
        
        # Definir a largura da coluna com base no conteúdo mais longo
        # O fator 1.2 ajuda a compensar a largura dos caracteres
        adjusted_width = max_length * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Congelar a linha de cabeçalho
    ws.freeze_panes = "A2"
    
    # Salvar e retornar
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="campaign_clicks.xlsx"'
    
    return response